import openpyxl
from openpyxl.styles import Font


async def export_to_excel(data, headings, filepath):
    """
    Exports data from PostgreSQL to an Excel spreadsheet using psycopg2.

    Arguments:
    connection - an open psycopg2 (this function does not close the connection)
    query_string - SQL to get data
    headings - list of strings to use as column headings
    filepath - path and filename of the Excel file

    psycopg2 and file handling errors bubble up to calling code.
    """

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.row_dimensions[1].font = Font(bold=True)

    # Spreadsheet row and column indexes start at 1,
    # so we use "start = 1" in enumerate, so
    # we don't need to add 1 to the indexes.
    for colno, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=colno).value = heading

    # This time we use "start = 2" to skip the heading row.
    for rowno, row in enumerate(data, start=2):
        for colno, cell_value in enumerate(row, start=1):
            sheet.cell(row=rowno, column=colno).value = cell_value

    wb.save(filepath)


import openpyxl
from openpyxl.styles import Font
from io import BytesIO

async def export_revenue_to_excel(data, headings, filepath):
    """
    Export revenue data to an Excel spreadsheet.
    Arguments:
    - data: List of revenue records to export.
    - headings: List of headings for the Excel columns.
    - filepath: Path and filename where the Excel file will be saved.
    """
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.row_dimensions[1].font = Font(bold=True)

    # Create the heading row in the Excel sheet.
    for colno, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=colno).value = heading

    # Fill in the data.
    for rowno, row in enumerate(data, start=2):
        for colno, cell_value in enumerate(row, start=1):
            sheet.cell(row=rowno, column=colno).value = cell_value

    # Save the workbook.
    wb.save(filepath)

